"""Write validated ETL updates to a copy of Programa.xlsx."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

from src.validators.validation_layer import ValidationResult

SHEET_NAME = "DIARIO"
COL_STATION = 2
COL_REGISTRATION = 4
COL_TICKETS = 5
COL_RESERVE_STATUS = 5
CELL_PROGRAM_DATE = "B3"
TEST_ROW = 68
RESERVE_HEADER_ROW = 69
TEST_TRAIN_COLUMN = 2
TEST_PKS_COLUMN = 3
TEST_REGISTRATION_COLUMN = 4
TEST_FIXED_TEXT_COLUMN = 5
TEST_START_TIME_COLUMN = 6
TEST_END_TIME_COLUMN = 7
TEST_DURATION_COLUMN = 8
TEST_DURATION_MERGE_END_COLUMN = 9
# A test consist may use either the operational slash or the word "y".
VALID_REGISTRATION_RE = re.compile(r"^[A-Z]\d{3}(?:(?:/|\s*Y\s*)[A-Z]\d{3})*$")
WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
MONTH_NAMES_ES = (
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
)


@dataclass(frozen=True)
class TestTrainInput:
    """Operator-entered details for one test train."""

    __test__ = False
    train: str
    pks: str
    registration: str
    start_time: str
    end_time: str


@dataclass(frozen=True)
class LoadResult:
    output_path: Path
    commercial_updates_written: int
    ticket_updates_written: int
    reserve_updates_written: int
    test_train_written: bool


class ExcelLoader:
    """Persist validated updates without changing Programa's base layout."""

    def load(
        self,
        validated_updates: ValidationResult,
        input_path: str | Path,
        output_path: str | Path,
        program_date: date | None = None,
        test_trains: list[TestTrainInput] | None = None,
    ) -> LoadResult:
        source = Path(input_path).resolve()
        destination = Path(output_path).resolve()
        if source == destination:
            raise ValueError("Output path must differ from input path; the template is never overwritten.")

        workbook = load_workbook(source, data_only=False)
        worksheet = workbook[SHEET_NAME]
        normalized_test_trains = self._normalize_test_trains(test_trains or [])

        if program_date is not None:
            self._write(worksheet, 3, 2, format_program_date(program_date))

        added_test_rows = self._prepare_test_rows(worksheet, len(normalized_test_trains))
        self._write_test_trains(worksheet, normalized_test_trains)

        self._synchronize_commercial_registration_merges(
            worksheet, validated_updates.commercial_updates
        )
        for update in validated_updates.commercial_updates:
            self._write(worksheet, update.target_row, COL_REGISTRATION, update.registration)
        for update in validated_updates.ticket_updates:
            self._write(worksheet, update.target_row, COL_TICKETS, update.tickets_sold)
        self._rebuild_reserve_section(
            worksheet,
            validated_updates.reserve_updates,
            reserve_header_row=RESERVE_HEADER_ROW + added_test_rows,
        )

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return LoadResult(
            output_path=destination,
            commercial_updates_written=len(validated_updates.commercial_updates),
            ticket_updates_written=len(validated_updates.ticket_updates),
            reserve_updates_written=len(validated_updates.reserve_updates),
            test_train_written=bool(normalized_test_trains),
        )

    @staticmethod
    def _normalize_test_trains(test_trains: list[TestTrainInput]) -> list[TestTrainInput]:
        normalized: list[TestTrainInput] = []
        for item in test_trains:
            if not all((item.train.strip(), item.pks.strip(), item.registration.strip(), item.start_time.strip(), item.end_time.strip())):
                raise ValueError("Every test train must include Tren, P.K.'s, MR, Hora Inicio and Hora Final.")
            registration = item.registration.strip().upper()
            if not VALID_REGISTRATION_RE.fullmatch(registration):
                raise ValueError(
                    "Test MR must use A000 or coupled registrations, for example "
                    "N001, R001/R006 or R001 y R006."
                )
            registration = re.sub(r"\s*Y\s*", " y ", registration)
            start_time = ExcelLoader._normalize_time(item.start_time, "Hora Inicio")
            end_time = ExcelLoader._normalize_time(item.end_time, "Hora Final")
            normalized.append(TestTrainInput(item.train.strip(), item.pks.strip(), registration, start_time, end_time))
        return normalized

    @staticmethod
    def _normalize_time(value: str, label: str) -> str:
        try:
            return datetime.strptime(value.strip(), "%H:%M").strftime("%H:%M")
        except ValueError as error:
            raise ValueError(f"{label} must use the 24-hour HH:MM format, for example 17:00.") from error

    @staticmethod
    def _duration_label(start_time: str, end_time: str) -> str:
        start = datetime.strptime(start_time, "%H:%M")
        end = datetime.strptime(end_time, "%H:%M")
        if end < start:
            end += timedelta(days=1)
        minutes = int((end - start).total_seconds() // 60)
        hours, remaining_minutes = divmod(minutes, 60)
        return f"{hours}h" if remaining_minutes == 0 else f"{hours}h {remaining_minutes:02d}m"

    @staticmethod
    def _prepare_test_rows(worksheet, test_count: int) -> int:
        extra_rows = max(0, test_count - 1)
        if extra_rows == 0:
            return 0

        affected_ranges = [
            (merged.min_row, merged.min_col, merged.max_row, merged.max_col)
            for merged in worksheet.merged_cells.ranges
            if merged.max_row >= RESERVE_HEADER_ROW
        ]
        for min_row, min_col, max_row, max_col in affected_ranges:
            worksheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        worksheet.insert_rows(RESERVE_HEADER_ROW, amount=extra_rows)
        for min_row, min_col, max_row, max_col in affected_ranges:
            if min_row >= RESERVE_HEADER_ROW:
                min_row += extra_rows
                max_row += extra_rows
            else:
                max_row += extra_rows
            worksheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

        for row in range(TEST_ROW + 1, TEST_ROW + test_count):
            worksheet.row_dimensions[row].height = worksheet.row_dimensions[TEST_ROW].height
            for column in range(1, worksheet.max_column + 1):
                worksheet.cell(row, column)._style = copy(worksheet.cell(TEST_ROW, column)._style)
            worksheet.merge_cells(
                start_row=row,
                start_column=TEST_DURATION_COLUMN,
                end_row=row,
                end_column=TEST_DURATION_MERGE_END_COLUMN,
            )
        return extra_rows

    @staticmethod
    def _write_test_trains(worksheet, test_trains: list[TestTrainInput]) -> None:
        for offset, item in enumerate(test_trains):
            row = TEST_ROW + offset
            ExcelLoader._write(worksheet, row, TEST_TRAIN_COLUMN, item.train)
            ExcelLoader._write(worksheet, row, TEST_PKS_COLUMN, item.pks)
            ExcelLoader._write(worksheet, row, TEST_REGISTRATION_COLUMN, item.registration)
            ExcelLoader._write(worksheet, row, TEST_FIXED_TEXT_COLUMN, "N/A")
            ExcelLoader._write(worksheet, row, TEST_START_TIME_COLUMN, item.start_time)
            ExcelLoader._write(worksheet, row, TEST_END_TIME_COLUMN, item.end_time)
            ExcelLoader._write(worksheet, row, TEST_DURATION_COLUMN, ExcelLoader._duration_label(item.start_time, item.end_time))
            for column in range(TEST_TRAIN_COLUMN, TEST_DURATION_COLUMN + 1):
                worksheet.cell(row, column).fill = WHITE_FILL

    @staticmethod
    def _synchronize_commercial_registration_merges(worksheet, commercial_updates) -> None:
        """Make Matrícula merges follow the PDF's combined-service rows.

        A single PDF row such as ``301-302`` means both services use one
        material unit.  Their cells are merged only if their physical service
        blocks in Programa are adjacent.  A non-adjacent pair must remain as
        two equal cells; merging it would cover unrelated services between
        them and corrupt the template's route layout.
        """

        if not commercial_updates:
            return

        test_header_row = next(
            (
                row
                for row in range(1, worksheet.max_row + 1)
                if worksheet.cell(row, COL_STATION).value == "Pruebas"
            ),
            None,
        )
        if test_header_row is None:
            raise ValueError("Programa template has no Pruebas section header.")
        commercial_end_row = test_header_row - 1

        service_starts = sorted(
            (update.target_row, update.service) for update in commercial_updates
        )
        service_spans: dict[str, tuple[int, int]] = {}
        for index, (start_row, service) in enumerate(service_starts):
            next_start = (
                service_starts[index + 1][0]
                if index + 1 < len(service_starts)
                else commercial_end_row + 1
            )
            service_spans[service] = (start_row, next_start - 1)

        desired_spans = set(service_spans.values())
        groups: dict[tuple[str, ...], list] = {}
        for update in commercial_updates:
            source_services = tuple(update.source_services or (update.service,))
            groups.setdefault(source_services, []).append(update)

        for source_services, group_updates in groups.items():
            if len(source_services) < 2 or len(group_updates) != len(source_services):
                continue
            try:
                spans = sorted(service_spans[service] for service in source_services)
            except KeyError:
                continue
            if all(spans[index][1] + 1 == spans[index + 1][0] for index in range(len(spans) - 1)):
                for span in spans:
                    desired_spans.discard(span)
                desired_spans.add((spans[0][0], spans[-1][1]))

        existing_merges = [
            (merged.min_row, merged.min_col, merged.max_row, merged.max_col)
            for merged in worksheet.merged_cells.ranges
            if merged.min_col == COL_REGISTRATION == merged.max_col
            and merged.max_row <= commercial_end_row
        ]
        # Only the top-left cell of an existing merge has the intended style.
        # Propagate that anchor style before unmerging, otherwise a newly
        # created sub-merge can inherit the blank child cell's default font.
        saved_styles = {
            row: copy(worksheet.cell(start_row, COL_REGISTRATION)._style)
            for start_row, _, end_row, _ in existing_merges
            for row in range(start_row, end_row + 1)
        }
        for start_row, start_column, end_row, end_column in existing_merges:
            worksheet.unmerge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=end_row,
                end_column=end_column,
            )
        for row, style in saved_styles.items():
            worksheet.cell(row, COL_REGISTRATION)._style = copy(style)
        for start_row, end_row in sorted(desired_spans):
            if end_row > start_row:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=COL_REGISTRATION,
                    end_row=end_row,
                    end_column=COL_REGISTRATION,
                )

    @staticmethod
    def _rebuild_reserve_section(worksheet, reserve_updates, reserve_header_row: int) -> None:
        """Resize and rebuild Reserva directly from the PDF's valid records.

        The PDF controls both the number of reserves and the stations where
        they are located.  The master template therefore supplies style, not
        a fixed capacity for each station.
        """

        reserve_start = reserve_header_row + 1
        reserve_end = reserve_start
        while reserve_end <= worksheet.max_row:
            values = [worksheet.cell(reserve_end, column).value for column in range(2, 10)]
            if all(value is None for value in values):
                break
            reserve_end += 1
        old_count = reserve_end - reserve_start
        if old_count <= 0:
            raise ValueError("Programa template has no rows below the Reserva header.")

        base_styles = {
            column: copy(worksheet.cell(reserve_start, column)._style)
            for column in range(1, worksheet.max_column + 1)
        }
        base_height = worksheet.row_dimensions[reserve_start].height
        new_count = len(reserve_updates)
        delta = new_count - old_count

        # openpyxl does not update merges when rows are inserted/deleted. Save
        # and restore every affected merge, except those in Reserva itself,
        # which will be recreated from the source records below.
        affected_ranges = [
            (merged.min_row, merged.min_col, merged.max_row, merged.max_col)
            for merged in worksheet.merged_cells.ranges
            if merged.max_row >= reserve_start
        ]
        for min_row, min_col, max_row, max_col in affected_ranges:
            worksheet.unmerge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

        if delta > 0:
            worksheet.insert_rows(reserve_end, amount=delta)
        elif delta < 0:
            worksheet.delete_rows(reserve_start + new_count, amount=-delta)

        for min_row, min_col, max_row, max_col in affected_ranges:
            # Old Reserva merges are deliberately replaced with dynamic groups.
            if min_row >= reserve_start and max_row < reserve_end:
                continue
            if min_row >= reserve_end:
                min_row += delta
                max_row += delta
            elif max_row >= reserve_end:
                max_row += delta
            worksheet.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

        for row in range(reserve_start, reserve_start + new_count):
            worksheet.row_dimensions[row].height = base_height
            for column, style in base_styles.items():
                cell = worksheet.cell(row, column)
                cell._style = copy(style)
                cell.value = None

        groups: dict[str, list] = {}
        for update in reserve_updates:
            groups.setdefault(update.workshop_station, []).append(update)

        current_row = reserve_start
        for station, updates in groups.items():
            group_end = current_row + len(updates) - 1
            worksheet.merge_cells(
                start_row=current_row,
                start_column=COL_STATION,
                end_row=group_end,
                end_column=3,
            )
            worksheet.cell(current_row, COL_STATION).value = station
            for update in updates:
                worksheet.merge_cells(
                    start_row=current_row,
                    start_column=COL_RESERVE_STATUS,
                    end_row=current_row,
                    end_column=9,
                )
                worksheet.cell(current_row, COL_REGISTRATION).value = update.registration
                worksheet.cell(current_row, COL_RESERVE_STATUS).value = (
                    "" if update.status.strip().upper() == "RESERVA" else update.status
                )
                current_row += 1

    @staticmethod
    def _write(worksheet, row: int, column: int, value: str | int) -> None:
        cell = worksheet.cell(row=row, column=column)
        if isinstance(cell, MergedCell):
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                    break
            else:
                raise ValueError(f"Merged cell {cell.coordinate} has no merge anchor.")
        cell.value = value


def format_program_date(value: date) -> str:
    return f"{value.day:02d} {MONTH_NAMES_ES[value.month - 1]}. {value.year}"
