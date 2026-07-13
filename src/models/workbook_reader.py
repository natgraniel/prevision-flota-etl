"""
workbook_reader.py

Reads the existing structure of `Programa.xlsx` (sheet `DIARIO`) so
that the Validation Layer can check extracted records against it, and
the Excel Loader can locate the correct anchor cells to write to.

This module does not validate business rules and does not write
anything to the workbook — it only reads and describes the workbook's
current shape. See docs/02_data_dictionary.md, sections 5 and 6, for
the section-marker convention and the "out of scope" row rule this
module implements.

Section detection: a section title is a row where columns B:I are
merged into a single cell (verified against the real workbook — see
docs/02_data_dictionary.md, section 5). Known section titles seen in
practice: "Circulaciones comerciales", "Pruebas", "Reserva". A section
ends at the first fully-empty row (all of B:I are None) — verified
against the real workbook, which has an empty row (85) immediately
after the last Reserve row (84), before an unrelated footnote
("Hora del centro.", a B:C merge, not a B:I section title) further
down. Closing the section on the first empty row correctly excludes
that footnote without needing to special-case it.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME = "DIARIO"
COL_CIRCULACION = 2  # B
COL_RUTA = 3  # C
COL_MATRICULA = 4  # D
COL_BOLETOS = 5  # E
SECTION_MERGE_MIN_COL = 2  # B
SECTION_MERGE_MAX_COL = 9  # I


@dataclass
class CommercialBlockRow:
    """One row within a commercial-service block in the workbook."""

    row: int
    service: str  # value of the Circulación anchor cell for this block, as text
    route: str | None
    is_anchor_row: bool  # True if this row is the merge anchor (top-left) for its block


@dataclass
class ReserveBlockRow:
    """One row within the Reserve section in the workbook."""

    row: int
    station: str | None
    is_anchor_row: bool


@dataclass
class WorkbookStructure:
    commercial_rows: list[CommercialBlockRow] = field(default_factory=list)
    reserve_rows: list[ReserveBlockRow] = field(default_factory=list)
    out_of_scope_rows: list[int] = field(default_factory=list)  # e.g. "Pruebas" section


def _is_section_title_row(ws: Worksheet, row: int) -> str | None:
    """Return the section title text if `row` is a B:I merged title row, else None."""
    for mc in ws.merged_cells.ranges:
        if (
            mc.min_row == row
            and mc.max_row == row
            and mc.min_col == SECTION_MERGE_MIN_COL
            and mc.max_col == SECTION_MERGE_MAX_COL
        ):
            return ws.cell(row=row, column=SECTION_MERGE_MIN_COL).value
    return None


def _merge_anchor_rows(ws: Worksheet, col: int) -> set[int]:
    """Return the set of anchor (top-left) rows for merges in a given column."""
    anchors = set()
    for mc in ws.merged_cells.ranges:
        if mc.min_col == col == mc.max_col:
            anchors.add(mc.min_row)
    return anchors


class WorkbookReader:
    """
    Reads docs/02_data_dictionary.md §5/§6-compliant structure from
    `Programa.xlsx`. Read-only — no validation, no writes.
    """

    def read(self, xlsx_path: str) -> WorkbookStructure:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb[SHEET_NAME]

        structure = WorkbookStructure()
        circulacion_anchors = _merge_anchor_rows(ws, COL_CIRCULACION)
        station_anchors = _merge_anchor_rows(ws, COL_CIRCULACION)

        current_section = None
        max_row = ws.max_row

        row = 1
        while row <= max_row:
            title = _is_section_title_row(ws, row)
            if title is not None:
                current_section = title
                consecutive_empty = 0
                row += 1
                continue

            row_values = [ws.cell(row=row, column=c).value for c in range(2, 10)]
            is_fully_empty = all(v is None for v in row_values)

            if is_fully_empty:
                current_section = None
                row += 1
                continue

            circulacion_val = ws.cell(row=row, column=COL_CIRCULACION).value

            # Skip the column-header row that immediately follows the
            # "Circulaciones comerciales" section title (verified: row 6
            # in the sample file contains the literal header labels).
            if circulacion_val == "Circulación":
                row += 1
                continue

            if current_section == "Circulaciones comerciales":
                route_val = ws.cell(row=row, column=COL_RUTA).value
                structure.commercial_rows.append(
                    CommercialBlockRow(
                        row=row,
                        service=str(circulacion_val) if circulacion_val is not None else "",
                        route=route_val,
                        is_anchor_row=row in circulacion_anchors,
                    )
                )
            elif current_section == "Reserva":
                station_val = ws.cell(row=row, column=COL_CIRCULACION).value
                structure.reserve_rows.append(
                    ReserveBlockRow(
                        row=row,
                        station=station_val,
                        is_anchor_row=row in station_anchors,
                    )
                )
            elif current_section == "Pruebas":
                structure.out_of_scope_rows.append(row)

            row += 1

        return structure


if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "data/raw/Programa_sample.xlsx"
    reader = WorkbookReader()
    structure = reader.read(path)

    print(f"Commercial rows: {len(structure.commercial_rows)}")
    for r in structure.commercial_rows:
        anchor_flag = "(anchor)" if r.is_anchor_row else ""
        print(f"  row={r.row:3d} service={r.service:8s} route={r.route!r} {anchor_flag}")

    print(f"\nReserve rows: {len(structure.reserve_rows)}")
    for r in structure.reserve_rows:
        anchor_flag = "(anchor)" if r.is_anchor_row else ""
        print(f"  row={r.row:3d} station={r.station!r} {anchor_flag}")

    print(f"\nOut-of-scope rows (e.g. Pruebas section): {structure.out_of_scope_rows}")
