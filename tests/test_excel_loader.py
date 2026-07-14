from datetime import date
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from src.extractors.pdf_extractor import PDFExtractor
from src.extractors.word_extractor import WordExtractor
from src.loaders.excel_loader import ExcelLoader, TestTrainInput
from src.models.workbook_reader import WorkbookReader
from src.transformers.transformation_layer import TransformationLayer
from src.validators.validation_layer import ValidationLayer


def test_loader_writes_multiple_test_trains_and_preserves_schedule(tmp_path: Path):
    raw = Path("data/raw")
    source = next(raw.glob("*.xlsx"))
    pdf = next(raw.glob("*.pdf"))
    docx = next(raw.glob("*.docx"))
    output = tmp_path / "Programa_actualizado.xlsx"
    source_schedule = load_workbook(source, data_only=False)["DIARIO"]["F7"].value

    pdf_result = PDFExtractor().extract(str(pdf))
    word_result = WordExtractor().extract(str(docx))
    transformed = TransformationLayer().transform(
        pdf_result.commercial_services, pdf_result.reserve, word_result.operations
    )
    # The source files vary daily. Simulate the valid PDF case where it lists
    # the physically adjacent template services in reverse order: 304-303.
    transformed.commercial_updates = [
        replace(
            update,
            registration="C004",
            source_services=("304", "303"),
        )
        if update.service in {"303", "304"}
        else update
        for update in transformed.commercial_updates
    ]
    validated = ValidationLayer().validate(transformed, WorkbookReader().read(str(source)))

    load_result = ExcelLoader().load(
        validated,
        source,
        output,
        program_date=date(2026, 7, 10),
        test_trains=[
            TestTrainInput("P009", "855+000 - 893+000", "R001 y R006", "17:00", "03:00"),
            TestTrainInput("P020", "893+000 - 900+000", "N001", "08:15", "10:45"),
        ],
    )
    worksheet = load_workbook(output, data_only=False)["DIARIO"]

    assert load_result.commercial_updates_written == 22
    assert load_result.ticket_updates_written == 32
    assert load_result.reserve_updates_written == 14
    assert load_result.test_train_written is True
    assert worksheet["D7"].value == "L003"
    merges = {str(merged) for merged in worksheet.merged_cells.ranges}
    # The source PDF combines 301-302, whose blocks are adjacent in Programa.
    assert "D27:D34" in merges
    # The PDF may list a pair in either order; the template's physical order
    # determines the vertical merge.
    assert "D35:D38" in merges
    assert "D39:D42" in merges
    # 601-604 share an MR in the PDF but are separated by 602-603 in Programa;
    # they must remain independent to avoid covering those intervening rows.
    assert "D55:D56" in merges
    assert "D61:D62" in merges
    # The commercial merge must end before the Pruebas section header.
    assert "B67:I67" in merges
    assert "D65:D66" in merges
    assert "D65:D67" not in merges
    assert worksheet["D35"].font.name == "Noto Sans"
    assert worksheet["D35"].font.bold is True
    assert worksheet["D35"].font.sz == 12
    assert worksheet["B3"].value == "10 Julio. 2026"
    assert worksheet["B68"].value == "P009"
    assert worksheet["C68"].value == "855+000 - 893+000"
    assert worksheet["D68"].value == "R001 y R006"
    assert worksheet["E68"].value == "N/A"
    assert worksheet["F68"].value == "17:00"
    assert worksheet["G68"].value == "03:00"
    assert worksheet["H68"].value == "10h"
    assert worksheet["B69"].value == "P020"
    assert worksheet["C69"].value == "893+000 - 900+000"
    assert worksheet["D69"].value == "N001"
    assert worksheet["E69"].value == "N/A"
    assert worksheet["F69"].value == "08:15"
    assert worksheet["G69"].value == "10:45"
    assert worksheet["H69"].value == "2h 30m"
    assert worksheet["B68"].fill.fgColor.rgb == "00FFFFFF"
    assert worksheet["H69"].fill.fgColor.rgb == "00FFFFFF"
    assert "H69:I69" in {str(merged) for merged in worksheet.merged_cells.ranges}
    assert worksheet["E7"].value == 134
    assert worksheet["F7"].value == source_schedule
    assert worksheet["B70"].value == "Reserva"
    assert worksheet["D71"].value == "L007"
    assert worksheet["E71"].value == "RESERVA EN ESTACION"
    assert worksheet["E73"].value is None
